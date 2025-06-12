from collections import namedtuple
from dataclasses import dataclass, field, fields
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from datacontract.export.exporter import Exporter
from datacontract.model.data_contract_specification import (
    Contact,
    DataContractSpecification,
    Info,
)

EXCEL_TEMPLATE_FILE_PATH = "datacontract/schemas/odcs-template.xlsx"
CellValue = namedtuple("CellValue", ["row_position", "column_position", "value"])


class ContractElements:
    def __init__(
        self, header_cell: CellValue, layout: str, cells: List[CellValue] = []
    ):
        self.header_cell = header_cell
        self.layout = layout
        self.cells = cells

    def __repr__(self):
        return f"ContractElements(header_cell={self.header_cell}, layout={self.layout})"


@dataclass
class SchemaBasicInfo:
    name: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 5, "Name"),
            layout="column",
            cells=[],
        )
    )
    type: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 6, "Type"),
            layout="column",
            cells=[],
        )
    )
    description: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 7, "Description"),
            layout="column",
            cells=[],
        )
    )
    business_name: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 8, "Business Name"),
            layout="column",
            cells=[],
        )
    )
    physical_name: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 9, "Physical Name"),
            layout="column",
            cells=[],
        )
    )
    data_granularity: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 10, "Data Granularity"),
            layout="column",
            cells=[],
        )
    )
    tags: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 11, "Tags"),
            layout="column",
            cells=[],
        )
    )


@dataclass
class SchemaFullInfo:
    property: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 13, "Property"),
            layout="column",
            cells=[],
        )
    )
    type: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 13, "Business Name"),
            layout="column",
            cells=[],
        )
    )
    logical_type: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(3, 13, "Logical Type"),
            layout="column",
            cells=[],
        )
    )
    physical_type: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(4, 13, "Physical Type"),
            layout="column",
            cells=[],
        )
    )
    examples: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(5, 13, "Example(s)"),
            layout="column",
            cells=[],
        )
    )
    description: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(6, 13, "Description"),
            layout="column",
            cells=[],
        )
    )
    required: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(7, 13, "Required"),
            layout="column",
            cells=[],
        )
    )
    unique: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(8, 13, "Unique"),
            layout="column",
            cells=[],
        )
    )
    classification: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(9, 13, "Classification"),
            layout="column",
            cells=[],
        )
    )
    tags: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(10, 13, "Tags"),
            layout="column",
            cells=[],
        )
    )
    quality_type: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(11, 13, "Quality Type"),
            layout="column",
            cells=[],
        )
    )
    quality_description: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(12, 13, "Quality Description"),
            layout="column",
            cells=[],
        )
    )
    authoritative_definition_url: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(13, 13, "Authoritative Definition URL"),
            layout="column",
            cells=[],
        )
    )
    authoritative_definition_type: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(14, 13, "Authoritative Definition Type"),
            layout="column",
            cells=[],
        )
    )
    physical_name: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(15, 13, "Physical Name"),
            layout="column",
            cells=[],
        )
    )
    primary_key: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(16, 13, "Primary Key"),
            layout="column",
            cells=[],
        )
    )
    primary_key_position: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(17, 13, "Primary Key Position"),
            layout="column",
            cells=[],
        )
    )
    partitioned: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(18, 13, "Partitioned"),
            layout="column",
            cells=[],
        )
    )
    partition_key_position: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(19, 13, "Partition Key Position"),
            layout="column",
            cells=[],
        )
    )
    encrypted_name: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(20, 13, "Encrypted Name"),
            layout="column",
            cells=[],
        )
    )
    transform_sources: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(21, 13, "Transform Sources"),
            layout="column",
            cells=[],
        )
    )
    transform_logic: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(22, 13, "Transform Logic"),
            layout="column",
            cells=[],
        )
    )
    transform_description: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(23, 13, "Transform Description"),
            layout="column",
            cells=[],
        )
    )
    critical_data_element_status: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(24, 13, "Critical Data Element Status"),
            layout="column",
            cells=[],
        )
    )
    maximum_items: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(25, 13, "Maximum Items"),
            layout="column",
            cells=[],
        )
    )
    minimum_items: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(26, 13, "Minimum Items"),
            layout="column",
            cells=[],
        )
    )
    unique_items: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(27, 13, "Unique Items"),
            layout="column",
            cells=[],
        )
    )
    format: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(28, 13, "Format"),
            layout="column",
            cells=[],
        )
    )
    minimum_length: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(29, 13, "Minimum Length"),
            layout="column",
            cells=[],
        )
    )
    maximum_length: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(30, 13, "Maximum Length"),
            layout="column",
            cells=[],
        )
    )
    exclusive_minimum: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(31, 13, "Exclusive Minimum"),
            layout="column",
            cells=[],
        )
    )
    minimum: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(32, 13, "Minimum"),
            layout="column",
            cells=[],
        )
    )
    exclusive_maximum: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(33, 13, "Exclusive Maximum"),
            layout="column",
            cells=[],
        )
    )
    maximum: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(34, 13, "Maximum"),
            layout="column",
            cells=[],
        )
    )
    multiple_of: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(35, 13, "Multiple Of"),
            layout="column",
            cells=[],
        )
    )
    minimum_properties: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(36, 13, "Minimum Properties"),
            layout="column",
            cells=[],
        )
    )
    maximum_properties: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(37, 13, "Maximum Properties"),
            layout="column",
            cells=[],
        )
    )
    required_properties: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(38, 13, "Required Properties"),
            layout="column",
            cells=[],
        )
    )
    pattern: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(39, 13, "Pattern"),
            layout="column",
            cells=[],
        )
    )


@dataclass
class Pricing:
    price_amount: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 4, "Price Amount"),
            layout="column",
            cells=[],
        )
    )
    price_currency: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 5, "Price Currency"),
            layout="column",
            cells=[],
        )
    )
    price_unit: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 6, "Price Unit"),
            layout="column",
            cells=[],
        )
    )


@dataclass
class Fundamentals:
    kind: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 4, "Kind"),
            layout="row",
            cells=[],
        )
    )

    api_version: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 5, "apiVersion"),
            layout="row",
            cells=[],
        )
    )
    id: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 7, "ID"),
            layout="row",
            cells=[],
        )
    )
    name: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 8, "Name"),
            layout="row",
            cells=[],
        )
    )
    version: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 9, "Version"),
            layout="row",
            cells=[],
        )
    )
    status: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 10, "Status"),
            layout="row",
            cells=[],
        )
    )
    owner: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 12, "Owner"),
            layout="row",
            cells=[],
        )
    )
    domain: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 14, "Domain"),
            layout="row",
            cells=[],
        )
    )
    data_product: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 15, "Data Product"),
            layout="row",
            cells=[],
        )
    )
    tenant: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 16, "Tenant"),
            layout="row",
            cells=[],
        )
    )
    description_purpose: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 19, "Purpose"),
            layout="row",
            cells=[],
        )
    )
    description_limitation: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 20, "Limitations"),
            layout="row",
            cells=[],
        )
    )
    description_usage: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 21, "Usage"),
            layout="row",
            cells=[],
        )
    )
    tags: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 23, "Tags"),
            layout="row",
            cells=[],
        )
    )


@dataclass
class Support:
    channel: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(1, 4, "Channel"),
            layout="column",
            cells=[],
        )
    )
    channel_url: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(2, 4, "Channel URL"),
            layout="column",
            cells=[],
        )
    )
    description: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(3, 4, "Description"),
            layout="column",
            cells=[],
        )
    )
    tool: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(4, 4, "Tool"),
            layout="column",
            cells=[],
        )
    )
    scope: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(5, 4, "Scope"),
            layout="column",
            cells=[],
        )
    )
    invitation_url: ContractElements = field(
        default_factory=lambda: ContractElements(
            header_cell=CellValue(6, 4, "Invitation URL"),
            layout="column",
            cells=[],
        )
    )


class ExcelExporter(Exporter):
    def export(
        self, data_contract, model, server, sql_server_type, export_args
    ) -> dict:
        return to_excel(data_contract)


def to_excel(data_contract_spec: DataContractSpecification):
    """
    Export data contract specification to Excel format.

    Args:
        data_contract_spec: DataContractSpecification object containing models to export
    """
    wb = openpyxl.load_workbook(EXCEL_TEMPLATE_FILE_PATH, data_only=True)

    for model_name, model in data_contract_spec.models.items():
        _create_schema_worksheet(wb, wb["Schema <table_name>"], model_name, model)

    info = getattr(data_contract_spec, "info", None)
    contact = getattr(info, "contact", None) if info else None
    _create_support_worksheet(wb, wb["Support"], contact)
    _create_fundamentals_worksheet(wb, wb["Fundamentals"], data_contract_spec.terms)
    wb.save("excel_export.xlsx")

    return wb


def _create_schema_worksheet(wb, ws_template_schema, model_name: str, model):
    """Create and populate a schema worksheet for a given model."""
    # Create new worksheet from template
    ws_schema = wb.copy_worksheet(ws_template_schema)
    ws_schema.title = f"Schema {model_name}"

    # Populate basic schema information
    _populate_schema_basic_info(ws_schema, model_name, model)

    # Populate detailed field information
    if model.fields:
        _populate_schema_detailed_info(ws_schema, model.fields)


def _create_support_worksheet(wb, ws_template_support, contact: Optional[Contact]):
    """Create and populate a support worksheet for a given model."""
    # Create new worksheet from template
    ws_support = ws_template_support
    ws_support.title = "Support"

    # Populate support information
    _populate_support_info(ws_support, contact)


def _populate_schema_basic_info(ws_schema, model_name: str, model):
    """Populate basic schema information in the worksheet."""
    schema_basic_info = SchemaBasicInfo()

    # Define basic info mappings
    basic_info_mappings = [
        (schema_basic_info.name, model_name),
        (schema_basic_info.type, _safe_getattr(model, "type")),
        (schema_basic_info.description, _safe_getattr(model, "description")),
        (schema_basic_info.business_name, _safe_getattr(model, "business_name")),
        (schema_basic_info.physical_name, _safe_getattr(model, "physical_name")),
        (schema_basic_info.data_granularity, _safe_getattr(model, "data_granularity")),
        (schema_basic_info.tags, ", ".join(getattr(model, "tags", []) or [])),
    ]

    # Populate cells
    for field_info, value in basic_info_mappings:
        field_info.cells.append(
            CellValue(
                field_info.header_cell.column_position,
                field_info.header_cell.row_position + 1,
                value,
            )
        )

    write_to_sheet(schema_basic_info, ws_schema)


def _populate_schema_detailed_info(ws_schema, model_fields):
    """Populate detailed field information in the worksheet."""
    schema_detailed_info = SchemaFullInfo()

    for counter, (field_name, field) in enumerate(model_fields.items(), start=1):
        _populate_schema_field_data(schema_detailed_info, field, counter)

    write_to_sheet(schema_detailed_info, ws_schema)


def _populate_schema_field_data(schema_detailed_info, field, counter: int):
    """Populate data for a single field across all detail categories."""
    # print(f"Populating field data for: {field} with counter {counter}")
    # Define field mappings with their corresponding values
    field_mappings = [
        (schema_detailed_info.property, _safe_getattr(field, "title")),
        (schema_detailed_info.type, _safe_getattr(field, "type")),
        (schema_detailed_info.logical_type, _safe_getattr(field, "logicalType")),
        (schema_detailed_info.physical_type, _safe_getattr(field, "physicalType")),
        (
            schema_detailed_info.examples,
            _format_list_value(getattr(field, "examples", [])),
        ),
        (schema_detailed_info.description, _safe_getattr(field, "description")),
        (schema_detailed_info.required, _safe_getattr(field, "required")),
        (schema_detailed_info.unique, _safe_getattr(field, "unique")),
        (schema_detailed_info.classification, _safe_getattr(field, "classification")),
        (schema_detailed_info.tags, _format_list_value(getattr(field, "tags", []))),
        (
            schema_detailed_info.quality_description,
            _safe_getattr(field, "qualityDescription"),
        ),
        (schema_detailed_info.quality_type, _safe_getattr(field, "qualityType")),
        (
            schema_detailed_info.authoritative_definition_url,
            _safe_getattr(field, "authoritativeDefinitionUrl"),
        ),
        (
            schema_detailed_info.authoritative_definition_type,
            _safe_getattr(field, "authoritativeDefinitionType"),
        ),
        (schema_detailed_info.physical_name, _safe_getattr(field, "physicalName")),
        (
            schema_detailed_info.primary_key,
            _format_list_value(getattr(field, "primary", [])),
        ),
        (
            schema_detailed_info.primary_key_position,
            _safe_getattr(field, "primaryKeyPosition"),
        ),
        (schema_detailed_info.partitioned, _safe_getattr(field, "partitioned")),
        (
            schema_detailed_info.partition_key_position,
            _safe_getattr(field, "partitionKeyPosition"),
        ),
        (
            schema_detailed_info.encrypted_name,
            _safe_str_getattr(field, "encryptedName"),
        ),
        (
            schema_detailed_info.transform_sources,
            _format_list_value(getattr(field, "transformSources", [])),
        ),
        (schema_detailed_info.transform_logic, _safe_getattr(field, "transformLogic")),
        (
            schema_detailed_info.transform_description,
            _safe_getattr(field, "transformDescription"),
        ),
        (
            schema_detailed_info.critical_data_element_status,
            _safe_str_getattr(field, "criticalDataElementStatus"),
        ),
        (schema_detailed_info.maximum_items, _safe_str_getattr(field, "maximumItems")),
        (schema_detailed_info.minimum_items, _safe_str_getattr(field, "minimumItems")),
        (schema_detailed_info.unique_items, _safe_str_getattr(field, "uniqueItems")),
        (schema_detailed_info.format, _safe_getattr(field, "format")),
        (
            schema_detailed_info.minimum_length,
            _safe_str_getattr(field, "minimumLength"),
        ),
        (
            schema_detailed_info.maximum_length,
            _safe_str_getattr(field, "maximumlength"),
        ),
        (
            schema_detailed_info.exclusive_minimum,
            _safe_str_getattr(field, "exclusiveMinimum"),
        ),
        (schema_detailed_info.minimum, _safe_str_getattr(field, "minimum")),
        (
            schema_detailed_info.exclusive_maximum,
            _safe_str_getattr(field, "exclusiveMaximum"),
        ),
        (schema_detailed_info.maximum, _safe_str_getattr(field, "maximum")),
        (schema_detailed_info.multiple_of, _safe_str_getattr(field, "multiple_of")),
        (
            schema_detailed_info.minimum_properties,
            _safe_str_getattr(field, "minimumProperties"),
        ),
        (
            schema_detailed_info.maximum_properties,
            _safe_str_getattr(field, "maximumProperties"),
        ),
        (
            schema_detailed_info.required_properties,
            "".join(getattr(field, "requiredProperties", []) or []),
        ),
        (schema_detailed_info.pattern, _safe_getattr(field, "pattern")),
    ]

    # Populate all field data
    for field_info, value in field_mappings:
        field_info.cells.append(
            CellValue(
                field_info.header_cell.column_position + counter,
                field_info.header_cell.row_position,
                value,
            )
        )


def _create_fundamentals_worksheet(wb, ws_fundamentals_schema, terms, make_copy=False):
    """Create and populate a fundamentals worksheet for a given model."""
    # Create new worksheet from template

    ws_fundamentals = (
        wb.copy_worksheet(ws_fundamentals_schema)
        if make_copy
        else ws_fundamentals_schema
    )
    ws_fundamentals = ws_fundamentals_schema
    ws_fundamentals.title = f"Fundamentals"

    # Populate fundamentals information
    _populate_fundamentals_info(ws_fundamentals, terms, 1)


def _populate_fundamentals_info(ws_fundamentals, terms, counter: int):
    """Populate fundamentals information in the worksheet."""
    fundamentals_info = Fundamentals()
    fundamentals_mappings = [
        (fundamentals_info.kind, getattr(terms, "kind", "")),
        (fundamentals_info.api_version, getattr(terms, "apiVersion", "")),
        (fundamentals_info.id, getattr(terms, "id", "")),
        (fundamentals_info.name, getattr(terms, "name", "")),
        (fundamentals_info.version, getattr(terms, "version", "")),
        (fundamentals_info.status, getattr(terms, "status", "")),
        (fundamentals_info.owner, getattr(terms, "owner", "")),
        (fundamentals_info.domain, getattr(terms, "domain", "")),
        (fundamentals_info.data_product, getattr(terms, "dataProduct", "")),
        (fundamentals_info.tenant, getattr(terms, "tenant", "")),
        (fundamentals_info.description_purpose, getattr(terms, "purpose", "")),
        (fundamentals_info.description_limitation, getattr(terms, "limitations", "")),
        (fundamentals_info.description_usage, getattr(terms, "usage", "")),
        (fundamentals_info.tags, ", ".join(getattr(terms, "tags", []) or [])),
    ]

    # Populate all field data
    for field_info, value in fundamentals_mappings:
        field_info.cells.append(
            CellValue(
                field_info.header_cell.column_position - 1,
                field_info.header_cell.row_position + counter,
                value,
            )
        )
        write_to_sheet(fundamentals_info, ws_fundamentals)


def _populate_support_info(ws_schema, info):
    """Populate detailed field information in the worksheet."""
    if not info:
        return

    support_info = Support()

    # for counter, (field_name, field) in enumerate(info, start=1):
    _populate_support_field_info(support_info, info, 1)

    write_to_sheet(support_info, ws_schema)


def _populate_support_field_info(support_field_info, info_field, counter: int) -> None:
    """
    Populate the support sheet from exactly one Contact object (or None).
    """

    mappings = [
        (support_field_info.channel, _safe_getattr(info_field, "name")),
        (support_field_info.channel_url, _safe_getattr(info_field, "url")),
        (support_field_info.description, _safe_getattr(info_field, "description")),
        (support_field_info.tool, _safe_getattr(info_field, "tool")),
        (support_field_info.scope, _safe_getattr(info_field, "scope")),
        (support_field_info.invitation_url, _safe_getattr(info_field, "invitationUrl")),
    ]

    for elem, val in mappings:
        elem.cells.append(
            CellValue(
                elem.header_cell.column_position + counter,
                elem.header_cell.row_position,
                val,
            )
        )


def _format_list_value(value_list) -> str:
    """Format a list value as a comma-separated string."""
    if not value_list:
        return ""
    return ", ".join(str(item) for item in value_list)


def _safe_getattr(obj, attr: str, default=""):
    """Safely get attribute with default value, handling None cases."""
    value = getattr(obj, attr, default)
    return value if value is not None else default


def _safe_str_getattr(obj, attr: str, default=""):
    """Safely get attribute as string with default value, handling None cases."""
    value = getattr(obj, attr, default)
    if value is None:
        return default
    return str(value)


def to_camel_case(text: str) -> str:
    words = text.split()
    if not words:
        return ""

    first_word = words[0].lower()
    rest = [word.capitalize() for word in words[1:]]
    return first_word + "".join(rest)


def write_to_sheet(sheet_mapped_obj, worksheet: Worksheet):
    """
    Writes data from a dataclass object to an Excel worksheet.
    Handles 'layout' (row/column) and 'span_cells'/'span_columns' for lists.
    Note: This version does not explicitly clear cells beyond the written data.
          If a new list is shorter than previous data in the Excel file,
          old values may persist in the remaining cells.
    """
    for field in fields(sheet_mapped_obj):
        # print(f"Processing field: {field.name} with layout: {field.metadata.get('layout', 'column')}")
        for contract_element_cell in getattr(sheet_mapped_obj, field.name).cells:
            row_position = contract_element_cell.row_position
            column_position = contract_element_cell.column_position
            print(
                f"Writing to cell at ({row_position}, {column_position}): {contract_element_cell.value}"
            )
            cell = worksheet.cell(row=row_position, column=column_position)
            cell.value = contract_element_cell.value
