import io
import logging
from decimal import Decimal
from typing import Any, List, Optional

import openpyxl
import requests
from open_data_contract_standard.model import (
    DataQuality,
    OpenDataContractStandard,
    SchemaObject,
    SchemaProperty,
)
from openpyxl.cell.cell import Cell
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from datacontract.export.exporter import Exporter

logger = logging.getLogger(__name__)

ODCS_EXCEL_TEMPLATE_URL = (
    "https://github.com/datacontract/open-data-contract-standard-excel-template/raw/refs/heads/main/odcs-template.xlsx"
)


class ExcelExporter(Exporter):
    """Excel exporter that uses the official ODCS template"""

    def __init__(self, export_format):
        super().__init__(export_format)

    def export(self, data_contract, schema_name, server, sql_server_type, export_args) -> bytes:
        """
        Export data contract to Excel using the official ODCS template

        Args:
            data_contract: OpenDataContractStandard to export
            model: Model name (not used for Excel export)
            server: Server name (not used for Excel export)
            sql_server_type: SQL server type (not used for Excel export)
            export_args: Additional export arguments (template can be specified here)

        Returns:
            Excel file as bytes
        """
        # The data_contract is now always ODCS
        odcs = data_contract

        # Get template from export_args if provided, otherwise use default
        template = export_args.get("template") if export_args else None
        return export_to_excel_bytes(odcs, template)


def export_to_excel_bytes(odcs: OpenDataContractStandard, template_path: Optional[str] = None) -> bytes:
    """
    Export ODCS to Excel format using the official template and return as bytes

    Args:
        odcs: OpenDataContractStandard object to export
        template_path: Optional path/URL to custom Excel template. If None, uses default template.

    Returns:
        Excel file as bytes
    """
    if template_path:
        workbook = create_workbook_from_template(template_path)
    else:
        workbook = create_workbook_from_template(ODCS_EXCEL_TEMPLATE_URL)

    try:
        fill_fundamentals(workbook, odcs)
        fill_schema(workbook, odcs)
        fill_quality(workbook, odcs)
        fill_custom_properties(workbook, odcs)
        fill_support(workbook, odcs)
        fill_team(workbook, odcs)
        fill_roles(workbook, odcs)
        fill_sla_properties(workbook, odcs)
        fill_servers(workbook, odcs)
        fill_pricing(workbook, odcs)

        # Set focus on the Fundamentals sheet
        workbook.active = workbook["Fundamentals"]

        # Force formula recalculation
        try:
            workbook.calculation.calcMode = "auto"
        except (AttributeError, ValueError):
            # Fallback for older openpyxl versions or if calcMode doesn't exist
            pass

        # Write to output stream
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    finally:
        workbook.close()


def create_workbook_from_template(template_path: str) -> Workbook:
    """Load Excel template from file path or URL"""
    try:
        # Convert Path object to string if needed
        template_path_str = str(template_path)
        logger.info(f"Processing template path: {template_path_str}")

        # Check if it's a URL
        if template_path_str.startswith(("http://", "https://")):
            logger.info(f"Identified as URL, downloading from: {template_path_str}")
            # Download from URL
            response = requests.get(template_path_str, timeout=30)
            response.raise_for_status()
            template_bytes = response.content
            workbook = openpyxl.load_workbook(io.BytesIO(template_bytes))
        else:
            logger.info(f"Identified as local file: {template_path_str}")
            # Load from local file
            workbook = openpyxl.load_workbook(template_path_str)

        return workbook
    except Exception as e:
        logger.error(f"Failed to load Excel template from {template_path}: {e}")
        raise RuntimeError(f"Failed to load Excel template: {e}")


def fill_fundamentals(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Fundamentals sheet with basic contract information"""
    set_cell_value_by_name(workbook, "apiVersion", odcs.apiVersion)
    set_cell_value_by_name(workbook, "kind", odcs.kind)
    set_cell_value_by_name(workbook, "id", odcs.id)
    set_cell_value_by_name(workbook, "name", odcs.name)
    set_cell_value_by_name(workbook, "version", odcs.version)
    set_cell_value_by_name(workbook, "status", odcs.status)
    set_cell_value_by_name(workbook, "domain", odcs.domain)
    set_cell_value_by_name(workbook, "dataProduct", odcs.dataProduct)
    set_cell_value_by_name(workbook, "tenant", odcs.tenant)

    # Set owner from custom properties
    owner_value = None
    if odcs.customProperties:
        for prop in odcs.customProperties:
            if prop.property == "owner":
                owner_value = prop.value
                break
    set_cell_value_by_name(workbook, "owner", owner_value)

    set_cell_value_by_name(workbook, "slaDefaultElement", odcs.slaDefaultElement)

    # Set description fields
    if odcs.description:
        set_cell_value_by_name(workbook, "description.purpose", odcs.description.purpose)
        set_cell_value_by_name(workbook, "description.limitations", odcs.description.limitations)
        set_cell_value_by_name(workbook, "description.usage", odcs.description.usage)

    # Set tags as comma-separated string
    if odcs.tags:
        set_cell_value_by_name(workbook, "tags", ",".join(odcs.tags))


def fill_pricing(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill pricing information"""
    if odcs.price:
        set_cell_value_by_name(workbook, "price.priceAmount", odcs.price.priceAmount)
        set_cell_value_by_name(workbook, "price.priceCurrency", odcs.price.priceCurrency)
        set_cell_value_by_name(workbook, "price.priceUnit", odcs.price.priceUnit)


def fill_schema(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill schema information by cloning template sheets"""
    # Get template sheet "Schema <table_name>"
    schema_template_sheet = workbook["Schema <table_name>"]

    if odcs.schema_:
        # Create copies for all schemas first
        new_sheets = []
        for schema in odcs.schema_:
            # Clone the template sheet
            new_sheet = workbook.copy_worksheet(schema_template_sheet)
            new_sheet.title = f"Schema {schema.name}"

            # Copy defined names with schema sheet scope to the new sheet
            copy_sheet_names(workbook, schema_template_sheet, new_sheet)

            # Move the new sheet before the template sheet
            schema_template_sheet_index = workbook.index(schema_template_sheet)
            new_sheet_index = workbook.index(new_sheet)

            workbook.move_sheet(new_sheet, offset=schema_template_sheet_index - new_sheet_index)

            new_sheets.append((new_sheet, schema))

        # Remove the template sheet before filling
        workbook.remove(schema_template_sheet)

        # Now fill in schema information for each copied sheet
        for new_sheet, schema in new_sheets:
            # Copy named ranges from template to new sheet (if needed)
            # Note: copy_worksheet should have copied the named ranges already

            # Fill in schema information
            fill_single_schema(new_sheet, schema)
    else:
        # Remove the template sheet even if no schemas
        workbook.remove(schema_template_sheet)


def copy_sheet_names(workbook: Workbook, template_sheet: Worksheet, new_sheet: Worksheet):
    """Copy worksheet-scoped named ranges from template sheet to new sheet"""
    try:
        # Copy worksheet-scoped defined names from template sheet to new sheet
        for name_str in template_sheet.defined_names:
            try:
                # Get the DefinedName object
                defined_name = template_sheet.defined_names[name_str]

                # Get the original range reference
                original_ref = defined_name.attr_text

                # Create new defined name with same name and reference but scoped to new sheet
                new_name = DefinedName(name_str, attr_text=original_ref.replace(template_sheet.title, new_sheet.title))

                # Add to the new sheet's defined names (worksheet-scoped)
                new_sheet.defined_names.add(new_name)

            except Exception as e:
                logger.warning(f"Failed to copy worksheet-scoped named range {name_str}: {e}")

    except Exception as e:
        logger.warning(f"Error copying sheet names: {e}")


def fill_single_schema(sheet: Worksheet, schema: SchemaObject):
    """Fill a single schema sheet with schema information using named ranges"""
    # Use worksheet-scoped named ranges that were copied from the template
    set_cell_value_by_name_in_sheet(sheet, "schema.name", schema.name)
    set_cell_value_by_name_in_sheet(
        sheet, "schema.physicalType", schema.physicalType if schema.physicalType else "table"
    )
    set_cell_value_by_name_in_sheet(sheet, "schema.description", schema.description)
    set_cell_value_by_name_in_sheet(sheet, "schema.businessName", schema.businessName)
    set_cell_value_by_name_in_sheet(sheet, "schema.physicalName", schema.physicalName)
    set_cell_value_by_name_in_sheet(sheet, "schema.dataGranularityDescription", schema.dataGranularityDescription)

    # Set tags as comma-separated string
    if schema.tags:
        set_cell_value_by_name_in_sheet(sheet, "schema.tags", ",".join(schema.tags))

    # Fill properties using the template's properties table structure
    if schema.properties:
        fill_properties_in_schema_sheet(sheet, schema.properties)


def fill_properties_in_schema_sheet(sheet: Worksheet, properties: List[SchemaProperty], prefix: str = ""):
    """Fill properties in the schema sheet using the template's existing properties table"""
    try:
        # The template already has a properties table starting at row 13 with headers
        # Find the header row and map column names to indices
        header_row_index = 13
        headers = get_headers_from_header_row(sheet, header_row_index)

        # Reverse the headers dict to map header_name -> column_index
        header_map = {header_name.lower(): col_idx for col_idx, header_name in headers.items()}

        # Fill properties starting after header row
        row_index = header_row_index + 1
        for property in properties:
            row_index = fill_single_property_template(sheet, row_index, prefix, property, header_map)

    except Exception as e:
        logger.warning(f"Error filling properties: {e}")


def fill_single_property_template(
    sheet: Worksheet, row_index: int, prefix: str, property: SchemaProperty, header_map: dict
) -> int:
    """Fill a single property row using the template's column structure"""
    property_name = f"{prefix}{'.' + property.name if property.name else ''}" if prefix else property.name

    # Helper function to set cell value by header name
    def set_by_header(header_name: str, value: Any):
        col_idx = header_map.get(header_name.lower())
        if col_idx is not None:
            sheet.cell(row=row_index, column=col_idx + 1).value = value

    # Fill property fields based on template headers
    set_by_header("Property", property_name)
    set_by_header("Business Name", property.businessName)
    set_by_header("Logical Type", property.logicalType)
    set_by_header("Physical Type", property.physicalType)
    set_by_header("Physical Name", property.physicalName)
    set_by_header("Description", property.description)
    set_by_header("Required", property.required)
    set_by_header("Unique", property.unique)
    set_by_header("Primary Key", property.primaryKey)
    set_by_header("Primary Key Position", property.primaryKeyPosition)
    set_by_header("Partitioned", property.partitioned)
    set_by_header("Partition Key Position", property.partitionKeyPosition)
    set_by_header("Classification", property.classification)
    set_by_header("Tags", ",".join(property.tags) if property.tags else "")
    set_by_header(
        "Example(s)", ",".join(map(str, property.examples)) if property.examples else ""
    )  # Note: using "Example(s)" as in template
    set_by_header("Encrypted Name", property.encryptedName)
    set_by_header(
        "Transform Sources", ",".join(property.transformSourceObjects) if property.transformSourceObjects else ""
    )
    set_by_header("Transform Logic", property.transformLogic)
    set_by_header("Critical Data Element Status", property.criticalDataElement)

    # Authoritative definitions
    if property.authoritativeDefinitions and len(property.authoritativeDefinitions) > 0:
        set_by_header("Authoritative Definition URL", property.authoritativeDefinitions[0].url)
        set_by_header("Authoritative Definition Type", property.authoritativeDefinitions[0].type)

    next_row_index = row_index + 1

    # Handle nested properties
    if property.properties:
        for nested_property in property.properties:
            next_row_index = fill_single_property_template(
                sheet, next_row_index, property_name, nested_property, header_map
            )

    # Handle array items
    if property.items:
        next_row_index = fill_single_property_template(
            sheet, next_row_index, f"{property_name}.items", property.items, header_map
        )

    return next_row_index


def fill_single_property_simple(
    sheet: Worksheet, row_index: int, prefix: str, property: SchemaProperty, header_map: dict = None
) -> int:
    """Fill a single property row using header names (deprecated - use fill_single_property_template instead)"""
    # This function is kept for backward compatibility but should use header_map if provided
    if header_map is None:
        # Fallback to the template-based approach
        header_row_index = 13
        headers = get_headers_from_header_row(sheet, header_row_index)
        header_map = {header_name.lower(): col_idx for col_idx, header_name in headers.items()}

    # Delegate to the template-based function
    return fill_single_property_template(sheet, row_index, prefix, property, header_map)


def fill_quality(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Quality sheet with quality data"""
    quality_sheet = workbook["Quality"]

    try:
        ref = name_to_ref(workbook, "quality")
        if not ref:
            logger.warning("No quality range found")
            return

        # Parse range to find header row
        header_row_index = parse_range_safely(ref)

        headers = get_headers_from_header_row(quality_sheet, header_row_index)
        current_row_index = header_row_index + 1

        # Iterate through all schemas
        if odcs.schema_:
            for schema in odcs.schema_:
                # Add schema-level quality attributes
                if schema.quality:
                    for quality in schema.quality:
                        row = get_or_create_row(quality_sheet, current_row_index)
                        fill_quality_row(row, headers, schema.name, None, quality)
                        current_row_index += 1

                # Add property-level quality attributes
                if schema.properties:
                    current_row_index = fill_properties_quality(
                        quality_sheet, headers, schema.name, schema.properties, current_row_index
                    )

    except Exception as e:
        logger.warning(f"Error filling quality: {e}")


def fill_properties_quality(
    sheet: Worksheet,
    headers: dict,
    schema_name: str,
    properties: List[SchemaProperty],
    start_row_index: int,
    prefix: str = "",
) -> int:
    """Recursively fill quality data for properties"""
    current_row_index = start_row_index

    for property in properties:
        if not property.name:
            continue

        full_property_name = f"{prefix}{'.' + property.name if property.name else ''}" if prefix else property.name

        # Add quality attributes for this property
        if property.quality:
            for quality in property.quality:
                row = get_or_create_row(sheet, current_row_index)
                fill_quality_row(row, headers, schema_name, full_property_name, quality)
                current_row_index += 1

        # Recursively handle nested properties
        if property.properties:
            current_row_index = fill_properties_quality(
                sheet, headers, schema_name, property.properties, current_row_index, full_property_name
            )

        # Handle array items
        if property.items:
            items_property_name = f"{full_property_name}.items"
            if property.items.quality:
                for quality in property.items.quality:
                    row = get_or_create_row(sheet, current_row_index)
                    fill_quality_row(row, headers, schema_name, items_property_name, quality)
                    current_row_index += 1

            # Handle nested properties in array items
            if property.items.properties:
                current_row_index = fill_properties_quality(
                    sheet, headers, schema_name, property.items.properties, current_row_index, items_property_name
                )

    return current_row_index


def fill_quality_row(row, headers: dict, schema_name: str, property_name: Optional[str], quality: DataQuality):
    """Fill a single quality row"""
    for cell_index, header_name in headers.items():
        header_lower = header_name.lower().strip()

        if header_lower == "schema":
            set_cell_value(row, cell_index, schema_name)
        elif header_lower == "property":
            set_cell_value(row, cell_index, property_name)
        elif header_lower == "quality type":
            set_cell_value(row, cell_index, quality.type)
        elif header_lower == "description":
            set_cell_value(row, cell_index, quality.description)
        elif header_lower == "rule (library)":
            set_cell_value(row, cell_index, quality.rule)
        elif header_lower == "query (sql)":
            set_cell_value(row, cell_index, quality.query)
        elif header_lower == "threshold operator":
            operator = get_threshold_operator(quality)
            set_cell_value(row, cell_index, operator)
        elif header_lower == "threshold value":
            value = get_threshold_value(quality)
            set_cell_value(row, cell_index, value)
        elif header_lower == "quality engine (custom)":
            set_cell_value(row, cell_index, quality.engine)
        elif header_lower == "implementation (custom)":
            set_cell_value(row, cell_index, quality.implementation)
        elif header_lower == "severity":
            set_cell_value(row, cell_index, quality.severity)
        elif header_lower == "scheduler":
            set_cell_value(row, cell_index, quality.scheduler)
        elif header_lower == "schedule":
            set_cell_value(row, cell_index, quality.schedule)


def get_threshold_operator(quality: DataQuality) -> Optional[str]:
    """Get the threshold operator from quality object"""
    if hasattr(quality, "mustBe") and quality.mustBe is not None:
        return "mustBe"
    elif hasattr(quality, "mustNotBe") and quality.mustNotBe is not None:
        return "mustNotBe"
    elif hasattr(quality, "mustBeGreaterThan") and quality.mustBeGreaterThan is not None:
        return "mustBeGreaterThan"
    elif hasattr(quality, "mustBeGreaterThanOrEqualTo") and quality.mustBeGreaterThanOrEqualTo is not None:
        return "mustBeGreaterThanOrEqualTo"
    elif hasattr(quality, "mustBeGreaterOrEqualTo") and quality.mustBeGreaterOrEqualTo is not None:
        return "mustBeGreaterOrEqualTo"
    elif hasattr(quality, "mustBeLessThan") and quality.mustBeLessThan is not None:
        return "mustBeLessThan"
    elif hasattr(quality, "mustBeLessThanOrEqualTo") and quality.mustBeLessThanOrEqualTo is not None:
        return "mustBeLessThanOrEqualTo"
    elif hasattr(quality, "mustBeLessOrEqualTo") and quality.mustBeLessOrEqualTo is not None:
        return "mustBeLessOrEqualTo" 
    elif hasattr(quality, "mustBeBetween") and quality.mustBeBetween is not None:
        return "mustBeBetween"
    elif hasattr(quality, "mustNotBeBetween") and quality.mustNotBeBetween is not None:
        return "mustNotBeBetween"
    return None


def get_threshold_value(quality: DataQuality) -> Optional[str]:
    """Get the threshold value from quality object"""
    if hasattr(quality, "mustBe") and quality.mustBe is not None:
        return str(quality.mustBe)
    elif hasattr(quality, "mustNotBe") and quality.mustNotBe is not None:
        return str(quality.mustNotBe)
    elif hasattr(quality, "mustBeGreaterThan") and quality.mustBeGreaterThan is not None:
        return str(quality.mustBeGreaterThan)
    elif hasattr(quality, "mustBeGreaterThanOrEqualTo") and quality.mustBeGreaterThanOrEqualTo is not None:
        return str(quality.mustBeGreaterThanOrEqualTo)
    elif hasattr(quality, "mustBeGreaterOrEqualTo") and quality.mustBeGreaterOrEqualTo is not None:
        return str(quality.mustBeGreaterOrEqualTo)
    elif hasattr(quality, "mustBeLessThan") and quality.mustBeLessThan is not None:
        return str(quality.mustBeLessThan)
    elif hasattr(quality, "mustBeLessThanOrEqualTo") and quality.mustBeLessThanOrEqualTo is not None:
        return str(quality.mustBeLessThanOrEqualTo)
    elif hasattr(quality, "mustBeLessOrEqualTo") and quality.mustBeLessOrEqualTo is not None:
        return str(quality.mustBeLessOrEqualTo)
    elif hasattr(quality, "mustBeBetween") and quality.mustBeBetween is not None and len(quality.mustBeBetween) >= 2:
        return f"[{quality.mustBeBetween[0]}, {quality.mustBeBetween[1]}]"
    elif (
        hasattr(quality, "mustNotBeBetween")
        and quality.mustNotBeBetween is not None
        and len(quality.mustNotBeBetween) >= 2
    ):
        return f"[{quality.mustNotBeBetween[0]}, {quality.mustNotBeBetween[1]}]"
    return None


def fill_custom_properties(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Custom Properties sheet"""
    try:
        ref = name_to_ref(workbook, "CustomProperties")
        if not ref:
            logger.warning("No CustomProperties range found")
            return

        custom_properties_sheet = workbook["Custom Properties"]

        # Parse range to find header row
        header_row_index = parse_range_safely(ref)

        # Fill custom properties excluding owner
        if odcs.customProperties:
            row_index = header_row_index + 1
            for prop in odcs.customProperties:
                if prop.property != "owner" and prop.property:
                    row = get_or_create_row(custom_properties_sheet, row_index)
                    set_cell_value(row, 0, prop.property)  # Property column
                    set_cell_value(row, 1, prop.value)  # Value column
                    row_index += 1

    except Exception as e:
        logger.warning(f"Error filling custom properties: {e}")


def fill_support(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Support sheet"""
    try:
        ref = name_to_ref(workbook, "support")
        if not ref:
            logger.warning("No support range found")
            return

        support_sheet = workbook["Support"]

        # Parse range to find header row
        header_row_index = parse_range_safely(ref)

        headers = get_headers_from_header_row(support_sheet, header_row_index)

        if odcs.support:
            for support_index, support_channel in enumerate(odcs.support):
                row = get_or_create_row(support_sheet, header_row_index + 1 + support_index)

                for cell_index, header_name in headers.items():
                    header_lower = header_name.lower()
                    if header_lower == "channel":
                        set_cell_value(row, cell_index, support_channel.channel)
                    elif header_lower == "channel url":
                        set_cell_value(row, cell_index, support_channel.url)
                    elif header_lower == "description":
                        set_cell_value(row, cell_index, support_channel.description)
                    elif header_lower == "tool":
                        set_cell_value(row, cell_index, support_channel.tool)
                    elif header_lower == "scope":
                        set_cell_value(row, cell_index, support_channel.scope)
                    elif header_lower == "invitation url":
                        set_cell_value(row, cell_index, support_channel.invitationUrl)

    except Exception as e:
        logger.warning(f"Error filling support: {e}")


def fill_team(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Team sheet"""
    try:
        ref = name_to_ref(workbook, "team")
        if not ref:
            logger.warning("No team range found")
            return

        team_sheet = workbook["Team"]

        # Parse range to find header row
        header_row_index = parse_range_safely(ref)

        headers = get_headers_from_header_row(team_sheet, header_row_index)

        if odcs.team:
            for team_index, team_member in enumerate(odcs.team):
                row = get_or_create_row(team_sheet, header_row_index + 1 + team_index)

                for cell_index, header_name in headers.items():
                    header_lower = header_name.lower()
                    if header_lower == "username":
                        set_cell_value(row, cell_index, team_member.username)
                    elif header_lower == "name":
                        set_cell_value(row, cell_index, team_member.name)
                    elif header_lower == "description":
                        set_cell_value(row, cell_index, team_member.description)
                    elif header_lower == "role":
                        set_cell_value(row, cell_index, team_member.role)
                    elif header_lower == "date in":
                        set_cell_value(row, cell_index, team_member.dateIn)
                    elif header_lower == "date out":
                        set_cell_value(row, cell_index, team_member.dateOut)
                    elif header_lower == "replaced by username":
                        set_cell_value(row, cell_index, team_member.replacedByUsername)

    except Exception as e:
        logger.warning(f"Error filling team: {e}")


def fill_roles(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Roles sheet using fixed table structure"""
    try:
        roles_sheet = workbook["Roles"]

        # From template analysis: Row 4 has headers
        header_row_index = 4
        headers = get_headers_from_header_row(roles_sheet, header_row_index)

        if odcs.roles:
            for role_index, role in enumerate(odcs.roles):
                row = get_or_create_row(roles_sheet, header_row_index + 1 + role_index)

                for cell_index, header_name in headers.items():
                    header_lower = header_name.lower()
                    if header_lower == "role":
                        set_cell_value(row, cell_index, role.role)
                    elif header_lower == "description":
                        set_cell_value(row, cell_index, role.description)
                    elif header_lower == "access":
                        set_cell_value(row, cell_index, role.access)
                    elif header_lower == "1st level approvers":
                        set_cell_value(row, cell_index, role.firstLevelApprovers)
                    elif header_lower == "2nd level approvers":
                        set_cell_value(row, cell_index, role.secondLevelApprovers)

    except Exception as e:
        logger.warning(f"Error filling roles: {e}")


def fill_sla_properties(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the SLA sheet using fixed table structure"""
    try:
        sla_sheet = workbook["SLA"]

        # From template analysis: Row 6 has the SLA properties table headers
        header_row_index = 6

        headers = get_headers_from_header_row(sla_sheet, header_row_index)

        if odcs.slaProperties:
            for sla_index, sla_prop in enumerate(odcs.slaProperties):
                row = get_or_create_row(sla_sheet, header_row_index + 1 + sla_index)

                for cell_index, header_name in headers.items():
                    header_lower = header_name.lower()
                    if header_lower == "property":
                        set_cell_value(row, cell_index, sla_prop.property)
                    elif header_lower == "value":
                        set_cell_value(row, cell_index, sla_prop.value)
                    elif header_lower == "extended value":
                        set_cell_value(row, cell_index, sla_prop.valueExt)
                    elif header_lower == "unit":
                        set_cell_value(row, cell_index, sla_prop.unit)
                    elif header_lower == "element":
                        set_cell_value(row, cell_index, sla_prop.element)
                    elif header_lower == "driver":
                        set_cell_value(row, cell_index, sla_prop.driver)

    except Exception as e:
        logger.warning(f"Error filling SLA properties: {e}")


def fill_servers(workbook: Workbook, odcs: OpenDataContractStandard):
    """Fill the Servers sheet"""
    try:
        servers_sheet = workbook["Servers"]

        if odcs.servers:
            for index, server in enumerate(odcs.servers):
                set_cell_value_by_column_index(servers_sheet, "servers.server", index, server.server)
                set_cell_value_by_column_index(servers_sheet, "servers.description", index, server.description)
                set_cell_value_by_column_index(servers_sheet, "servers.environment", index, server.environment)
                set_cell_value_by_column_index(servers_sheet, "servers.type", index, server.type)

                # Type-specific fields
                server_type = server.type
                if server_type == "azure":
                    set_cell_value_by_column_index(servers_sheet, "servers.azure.location", index, server.location)
                    set_cell_value_by_column_index(servers_sheet, "servers.azure.format", index, server.format)
                    set_cell_value_by_column_index(servers_sheet, "servers.azure.delimiter", index, server.delimiter)
                elif server_type == "bigquery":
                    set_cell_value_by_column_index(servers_sheet, "servers.bigquery.project", index, server.project)
                    set_cell_value_by_column_index(servers_sheet, "servers.bigquery.dataset", index, server.dataset)
                elif server_type == "databricks":
                    set_cell_value_by_column_index(servers_sheet, "servers.databricks.catalog", index, server.catalog)
                    set_cell_value_by_column_index(servers_sheet, "servers.databricks.host", index, server.host)
                    set_cell_value_by_column_index(servers_sheet, "servers.databricks.schema", index, server.schema_)
                elif server_type == "glue":
                    set_cell_value_by_column_index(servers_sheet, "servers.glue.account", index, server.account)
                    set_cell_value_by_column_index(servers_sheet, "servers.glue.database", index, server.database)
                    set_cell_value_by_column_index(servers_sheet, "servers.glue.format", index, server.format)
                    set_cell_value_by_column_index(servers_sheet, "servers.glue.location", index, server.location)
                elif server_type == "kafka":
                    set_cell_value_by_column_index(servers_sheet, "servers.kafka.format", index, server.format)
                    set_cell_value_by_column_index(servers_sheet, "servers.kafka.host", index, server.host)
                elif server_type == "oracle":
                    set_cell_value_by_column_index(servers_sheet, "servers.oracle.host", index, server.host)
                    set_cell_value_by_column_index(servers_sheet, "servers.oracle.port", index, server.port)
                    set_cell_value_by_column_index(servers_sheet, "servers.oracle.servicename", index, server.serviceName)
                elif server_type == "postgres":
                    set_cell_value_by_column_index(servers_sheet, "servers.postgres.database", index, server.database)
                    set_cell_value_by_column_index(servers_sheet, "servers.postgres.host", index, server.host)
                    set_cell_value_by_column_index(servers_sheet, "servers.postgres.port", index, server.port)
                    set_cell_value_by_column_index(servers_sheet, "servers.postgres.schema", index, server.schema_)
                elif server_type == "s3":
                    set_cell_value_by_column_index(servers_sheet, "servers.s3.delimiter", index, server.delimiter)
                    set_cell_value_by_column_index(servers_sheet, "servers.s3.endpointUrl", index, server.endpointUrl)
                    set_cell_value_by_column_index(servers_sheet, "servers.s3.format", index, server.format)
                    set_cell_value_by_column_index(servers_sheet, "servers.s3.location", index, server.location)
                elif server_type == "snowflake":
                    set_cell_value_by_column_index(servers_sheet, "servers.snowflake.account", index, server.account)
                    set_cell_value_by_column_index(servers_sheet, "servers.snowflake.database", index, server.database)
                    set_cell_value_by_column_index(servers_sheet, "servers.snowflake.host", index, server.host)
                    set_cell_value_by_column_index(servers_sheet, "servers.snowflake.port", index, server.port)
                    set_cell_value_by_column_index(servers_sheet, "servers.snowflake.schema", index, server.schema_)
                    set_cell_value_by_column_index(servers_sheet, "servers.snowflake.warehouse", index, server.warehouse)
                elif server_type == "sqlserver":
                    set_cell_value_by_column_index(servers_sheet, "servers.sqlserver.database", index, server.database)
                    set_cell_value_by_column_index(servers_sheet, "servers.sqlserver.host", index, server.host)
                    set_cell_value_by_column_index(servers_sheet, "servers.sqlserver.port", index, server.port)
                    set_cell_value_by_column_index(servers_sheet, "servers.sqlserver.schema", index, server.schema_)
                else:
                    # Custom/unknown server type - export all possible fields
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.account", index, server.account)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.catalog", index, server.catalog)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.database", index, server.database)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.dataset", index, server.dataset)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.delimiter", index, server.delimiter)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.endpointUrl", index, server.endpointUrl)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.format", index, server.format)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.host", index, server.host)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.location", index, server.location)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.path", index, server.path)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.port", index, server.port)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.project", index, server.project)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.schema", index, server.schema_)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.serviceName", index, server.serviceName)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.stagingDir", index, server.stagingDir)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.warehouse", index, server.warehouse)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.region", index, server.region)
                    set_cell_value_by_column_index(servers_sheet, "servers.custom.regionName", index, server.regionName)

    except Exception as e:
        logger.warning(f"Error filling servers: {e}")


# Helper functions


def find_cell_by_name(workbook: Workbook, name: str) -> Optional[Cell]:
    """Find a cell by its named range"""
    try:
        ref = name_to_ref(workbook, name)
        if not ref:
            return None
        return find_cell_by_ref(workbook, ref)
    except Exception:
        return None


def find_cell_by_name_in_sheet(sheet: Worksheet, name: str) -> Optional[Cell]:
    """Find a cell by its named range within a specific sheet"""
    try:
        # Access worksheet-scoped defined names directly
        for named_range in sheet.defined_names:
            if named_range == name:
                destinations = sheet.defined_names[named_range].destinations
                for sheet_title, coordinate in destinations:
                    if sheet_title == sheet.title:
                        return sheet[coordinate]
    except Exception:
        return None
    return None


def find_cell_by_ref(workbook: Workbook, cell_ref: str) -> Optional[Cell]:
    """Find a cell by its reference"""
    try:
        from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

        # Parse the reference
        if "!" in cell_ref:
            sheet_name, coord = cell_ref.split("!")
            sheet_name = sheet_name.strip("'")
            sheet = workbook[sheet_name]
        else:
            coord = cell_ref
            sheet = workbook.active

        # Remove $ signs
        coord = coord.replace("$", "")
        col_letter, row_num = coordinate_from_string(coord)
        col_num = column_index_from_string(col_letter)

        return sheet.cell(row=int(row_num), column=col_num)
    except Exception:
        return None


def find_cell_by_ref_in_sheet(sheet: Worksheet, cell_ref: str) -> Optional[Cell]:
    """Find a cell by its reference within a specific sheet"""
    try:
        from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

        # Remove sheet name if present
        if "!" in cell_ref:
            _, coord = cell_ref.split("!")
        else:
            coord = cell_ref

        # Remove $ signs
        coord = coord.replace("$", "")
        col_letter, row_num = coordinate_from_string(coord)
        col_num = column_index_from_string(col_letter)

        return sheet.cell(row=int(row_num), column=col_num)
    except Exception:
        return None


def name_to_ref(workbook: Workbook, name: str) -> Optional[str]:
    """Get the reference for a named range in the workbook"""
    try:
        defined_name = workbook.defined_names.get(name)
        if defined_name:
            return defined_name.attr_text
    except Exception:
        pass
    return None


def name_to_ref_in_sheet(sheet: Worksheet, name: str) -> Optional[str]:
    """Get the reference for a named range in a specific sheet"""
    try:
        workbook = sheet.parent
        defined_names = [dn for dn in workbook.defined_names if dn.name == name]
        for dn in defined_names:
            if sheet.title in dn.attr_text:
                return dn.attr_text
    except Exception:
        pass
    return None


def set_cell_value_by_name(workbook: Workbook, cell_name: str, value: Any):
    """Set cell value by named range"""
    cell = find_cell_by_name(workbook, cell_name)
    if cell:
        set_cell_value_direct(cell, value)
    else:
        logger.warning(f"Cell with name {cell_name} not found in workbook")


def set_cell_value_by_name_in_sheet(sheet: Worksheet, cell_name: str, value: Any):
    """Set cell value by named range within a specific sheet"""
    cell = find_cell_by_name_in_sheet(sheet, cell_name)
    if cell:
        set_cell_value_direct(cell, value)
    else:
        logger.warning(f"Cell with name {cell_name} not found in sheet {sheet.title}")


def set_cell_value_by_column_index(sheet: Worksheet, name: str, column_index: int, value: Any):
    """Set cell value by column offset from named range"""
    try:
        workbook = sheet.parent
        first_cell = find_cell_by_name(workbook, name)
        if first_cell:
            target_cell = sheet.cell(row=first_cell.row, column=first_cell.column + column_index)
            set_cell_value_direct(target_cell, value)
    except Exception as e:
        logger.warning(f"Error setting cell value by column index: {e}")


def set_cell_value_direct(cell: Cell, value: Any):
    """Set cell value directly"""
    if value is not None:
        if isinstance(value, bool):
            cell.value = value
        elif isinstance(value, (int, float, Decimal)):
            cell.value = float(value)
        else:
            cell.value = str(value)
    else:
        cell.value = None


def set_cell_value(row, cell_index: int, value: Any):
    """Set cell value in a row at specific index"""
    cell = get_or_create_cell(row, cell_index)
    set_cell_value_direct(cell, value)


def get_or_create_row(sheet: Worksheet, row_index: int):
    """Get or create a row at the specified index"""
    try:
        return sheet[row_index]
    except (IndexError, KeyError):
        # If row doesn't exist, create it
        while len(list(sheet.rows)) < row_index:
            sheet.append([])
        return sheet[row_index]


def get_or_create_cell(row, cell_index: int) -> Cell:
    """Get or create a cell at the specified index in a row"""
    try:
        return row[cell_index]
    except IndexError:
        # Extend the row if needed
        while len(row) <= cell_index:
            row.append(None)
        return row[cell_index]


def parse_range_safely(ref: str) -> int:
    """Parse a range reference and return the starting row number"""
    try:
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(ref)
        return min_row
    except Exception:
        # Handle malformed ranges - extract row number from range like "Quality!$A$4:$AZ$300"
        if ":" in ref:
            start_ref = ref.split(":")[0]
            if "!" in start_ref:
                start_ref = start_ref.split("!")[-1]
            start_ref = start_ref.replace("$", "")
            # Extract row number
            import re

            row_match = re.search(r"(\d+)", start_ref)
            if row_match:
                return int(row_match.group(1))
        return 1


def get_headers_from_header_row(sheet: Worksheet, header_row_index: int) -> dict:
    """Get headers from a row and return as dict mapping cell_index -> header_name"""
    headers = {}
    try:
        header_row = sheet[header_row_index]
        for cell_index, cell in enumerate(header_row):
            if cell.value:
                headers[cell_index] = str(cell.value).strip()
    except Exception as e:
        logger.warning(f"Error getting headers from row {header_row_index}: {e}")
    return headers
