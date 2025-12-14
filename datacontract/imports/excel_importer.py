import logging
import os
from decimal import Decimal
from typing import Any, Dict, List, Optional

import openpyxl
from open_data_contract_standard.model import (
    AuthoritativeDefinition,
    CustomProperty,
    DataQuality,
    OpenDataContractStandard,
    Role,
    SchemaObject,
    SchemaProperty,
    Server,
    ServiceLevelAgreementProperty,
    Support,
    Team,
    TeamMember,
)
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from datacontract.imports.importer import Importer
from datacontract.model.exceptions import DataContractException

logger = logging.getLogger(__name__)


class ExcelImporter(Importer):
    def import_source(
        self,
        source: str,
        import_args: dict,
    ) -> OpenDataContractStandard:
        return import_excel_as_odcs(source)


def import_excel_as_odcs(excel_file_path: str) -> OpenDataContractStandard:
    """
    Import an Excel file and convert it to an OpenDataContractStandard object

    Args:
        excel_file_path: Path to the Excel file

    Returns:
        OpenDataContractStandard object
    """
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    try:
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    except Exception as e:
        raise DataContractException(
            type="schema",
            name="Parse excel contract",
            reason=f"Failed to open Excel file: {excel_file_path}",
            engine="datacontract",
            original_exception=e,
        )

    try:
        # Get description values
        purpose = get_cell_value_by_name(workbook, "description.purpose")
        limitations = get_cell_value_by_name(workbook, "description.limitations")
        usage = get_cell_value_by_name(workbook, "description.usage")

        # Build description dict
        description = None
        if purpose or limitations or usage:
            description = {"purpose": purpose, "limitations": limitations, "usage": usage}

        # Get tags as a list
        tags_str = get_cell_value_by_name(workbook, "tags")
        tags = None
        if tags_str:
            tags = [tag.strip() for tag in tags_str.split(",") if tag.strip()]

        # Import quality data first (standalone from schemas)
        quality_map = import_quality(workbook)

        # Import schemas
        schemas = import_schemas(workbook)

        # Attach quality to schemas and properties
        schemas_with_quality = attach_quality_to_schemas(schemas, quality_map)

        # Import other components
        support = import_support(workbook)
        team = import_team(workbook)
        roles = import_roles(workbook)
        sla_properties = import_sla_properties(workbook)
        servers = import_servers(workbook)
        price = import_price(workbook)
        custom_properties = import_custom_properties(workbook)

        # Create the ODCS object with proper object creation
        odcs = OpenDataContractStandard(
            apiVersion=get_cell_value_by_name(workbook, "apiVersion"),
            kind=get_cell_value_by_name(workbook, "kind"),
            id=get_cell_value_by_name(workbook, "id"),
            name=get_cell_value_by_name(workbook, "name"),
            version=get_cell_value_by_name(workbook, "version"),
            status=get_cell_value_by_name(workbook, "status"),
            domain=get_cell_value_by_name(workbook, "domain"),
            dataProduct=get_cell_value_by_name(workbook, "dataProduct"),
            tenant=get_cell_value_by_name(workbook, "tenant"),
            description=description,
            tags=tags,
            schema=schemas_with_quality,
            support=support,
            price=price,
            team=team,
            roles=roles,
            slaDefaultElement=get_cell_value_by_name(workbook, "slaDefaultElement"),
            slaProperties=sla_properties,
            servers=servers,
            customProperties=custom_properties,
        )

        return odcs
    except Exception as e:
        logger.error(f"Error importing Excel file: {str(e)}")
        raise DataContractException(
            type="schema",
            name="Parse excel contract",
            reason=f"Failed to parse Excel file: {excel_file_path}",
            engine="datacontract",
            original_exception=e,
        )
    finally:
        workbook.close()


def import_schemas(workbook) -> Optional[List[SchemaObject]]:
    """Extract schema information from sheets starting with 'Schema '"""
    schemas = []

    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("Schema ") and sheet_name != "Schema <table_name>":
            sheet = workbook[sheet_name]
            schema_name = get_cell_value_by_name_in_sheet(sheet, "schema.name")

            if not schema_name:
                continue

            schema = SchemaObject(
                name=schema_name,
                logicalType="object",
                physicalType=get_cell_value_by_name_in_sheet(sheet, "schema.physicalType"),
                physicalName=get_cell_value_by_name_in_sheet(sheet, "schema.physicalName"),
                description=get_cell_value_by_name_in_sheet(sheet, "schema.description"),
                businessName=get_cell_value_by_name_in_sheet(sheet, "schema.businessName"),
                dataGranularityDescription=get_cell_value_by_name_in_sheet(sheet, "schema.dataGranularityDescription"),
                authoritativeDefinitions=None,
                properties=import_properties(sheet),
                quality=None,  # Quality will be attached later
                customProperties=None,
                tags=None,
            )

            # Get tags
            tags_str = get_cell_value_by_name_in_sheet(sheet, "schema.tags")
            if tags_str:
                schema.tags = [tag.strip() for tag in tags_str.split(",") if tag.strip()]

            schemas.append(schema)

    return schemas if schemas else None


def import_properties(sheet) -> Optional[List[SchemaProperty]]:
    """Extract properties from the schema sheet"""
    try:
        # Find the properties table
        properties_range = get_range_by_name_in_sheet(sheet, "schema.properties")
        if not properties_range:
            return None

        # Get header row to map column names to indices
        header_row = list(sheet.rows)[properties_range[0] - 1]  # Convert to 0-based indexing
        headers = {}
        for i, cell in enumerate(header_row):
            if cell.value:
                headers[cell.value.lower()] = i

        # Process property rows
        property_lookup = {}  # Dictionary to keep track of properties by name for nesting

        # First, create all properties
        for row_idx in range(properties_range[0], properties_range[1]):
            if len(list(sheet.rows)) < row_idx + 1:
                break
            row = list(sheet.rows)[row_idx]

            # Skip empty rows or header row
            property_name = get_cell_value(row, headers.get("property"))
            if not property_name or row_idx == properties_range[0] - 1:
                continue

            # Create property object
            property_obj = SchemaProperty(
                name=property_name,
                logicalType=get_cell_value(row, headers.get("logical type")),
                logicalTypeOptions=import_logical_type_options(row, headers),
                physicalType=get_cell_value(row, headers.get("physical type")),
                physicalName=get_cell_value(row, headers.get("physical name")),
                description=get_cell_value(row, headers.get("description")),
                businessName=get_cell_value(row, headers.get("business name")),
                required=parse_boolean(get_cell_value(row, headers.get("required"))),
                unique=parse_boolean(get_cell_value(row, headers.get("unique"))),
                primaryKey=parse_boolean(get_cell_value(row, headers.get("primary key"))),
                primaryKeyPosition=parse_integer(get_cell_value(row, headers.get("primary key position"))),
                partitioned=parse_boolean(get_cell_value(row, headers.get("partitioned"))),
                partitionKeyPosition=parse_integer(get_cell_value(row, headers.get("partition key position"))),
                criticalDataElement=parse_boolean(get_cell_value(row, headers.get("critical data element status"))),
                classification=get_cell_value(row, headers.get("classification")),
                transformLogic=get_cell_value(row, headers.get("transform logic")),
                transformDescription=get_cell_value(row, headers.get("transform description")),
                encryptedName=get_cell_value(row, headers.get("encrypted name")),
                properties=None,
                items=None,
                tags=get_property_tags(headers, row),
            )

            # Authoritative definitions
            authoritative_definition_url = get_cell_value(row, headers.get("authoritative definition url"))
            authoritative_definition_type = get_cell_value(row, headers.get("authoritative definition type"))
            if authoritative_definition_url and authoritative_definition_type:
                property_obj.authoritativeDefinitions = [
                    AuthoritativeDefinition(
                        url=authoritative_definition_url,
                        type=authoritative_definition_type,
                    )
                ]

            # Quality will be attached later via quality_map
            property_obj.quality = None

            # Transform sources
            transform_sources = get_cell_value(row, headers.get("transform sources"))
            if transform_sources:
                property_obj.transformSourceObjects = [
                    src.strip() for src in transform_sources.split(",") if src.strip()
                ]

            # Examples
            examples = get_cell_value(row, headers.get("example(s)"))
            if examples:
                property_obj.examples = [ex.strip() for ex in examples.split(",") if ex.strip()]

            # Add to lookup dictionary
            property_lookup[property_name] = property_obj

        # Now organize nested properties
        root_properties = []
        for name, prop in property_lookup.items():
            if "." in name:
                # This is a nested property
                parent_name = name.rsplit(".", 1)[0]
                child_name = name.rsplit(".", 1)[1]

                if parent_name in property_lookup:
                    parent_prop = property_lookup[parent_name]
                    # Update the property name to be just the child part
                    prop.name = child_name

                    # If parent is an array, set as items
                    if parent_prop.logicalType == "array":
                        parent_prop.items = prop
                    else:
                        # Otherwise add to properties list
                        if parent_prop.properties is None:
                            parent_prop.properties = []
                        parent_prop.properties.append(prop)
            else:
                # This is a root property
                root_properties.append(prop)

        return root_properties if root_properties else None
    except Exception as e:
        logger.warning(f"Error importing properties: {str(e)}")
        return None


def import_logical_type_options(row, headers):
    """Import logical type options from property row"""

    required_props = get_cell_value(row, headers.get("required properties"))

    required_props_list = None
    if required_props:
        required_props_list = [prop.strip() for prop in required_props.split(",") if prop.strip()]

    logical_type_options_dict = {
        "minLength": parse_integer(get_cell_value(row, headers.get("minimum length"))),
        "maxLength": parse_integer(get_cell_value(row, headers.get("maximum length"))),
        "pattern": get_cell_value(row, headers.get("pattern")),
        "format": get_cell_value(row, headers.get("format")),
        "exclusiveMaximum": parse_boolean(get_cell_value(row, headers.get("exclusive maximum"))),
        "exclusiveMinimum": parse_boolean(get_cell_value(row, headers.get("exclusive minimum"))),
        "minimum": get_cell_value(row, headers.get("minimum")),
        "maximum": get_cell_value(row, headers.get("maximum")),
        "multipleOf": get_cell_value(row, headers.get("multiple of")),
        "minItems": parse_integer(get_cell_value(row, headers.get("minimum items"))),
        "maxItems": parse_integer(get_cell_value(row, headers.get("maximum items"))),
        "uniqueItems": parse_boolean(get_cell_value(row, headers.get("unique items"))),
        "maxProperties": parse_integer(get_cell_value(row, headers.get("maximum properties"))),
        "minProperties": parse_integer(get_cell_value(row, headers.get("minimum properties"))),
        "required": required_props_list,
    }

    for dict_key in list(logical_type_options_dict.keys()):
        if logical_type_options_dict[dict_key] is None:
            del logical_type_options_dict[dict_key]

    if len(logical_type_options_dict) == 0:
        return None
    return logical_type_options_dict


def get_property_tags(headers, row):
    tags_value = get_cell_value(row, headers.get("tags"))
    if tags_value:
        return [tag.strip() for tag in tags_value.split(",") if tag.strip()]
    return None


def parse_boolean(value):
    """Parse a string value to boolean"""
    if value is None:
        return None
    value = value.lower().strip()
    return value == "true" or value == "yes" or value == "1"


def parse_integer(value):
    """Parse a string value to integer"""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def get_range_by_name_in_workbook(workbook: Workbook, name: str) -> tuple | None:
    """Find the range (start_row, end_row) of a named range in a workbook"""
    try:
        for named_range in workbook.defined_names:
            if named_range == name:
                destinations = workbook.defined_names[named_range].destinations
                for sheet_title, range_address in destinations:
                    if ":" in range_address:
                        # Convert Excel range to row numbers
                        start_ref, end_ref = range_address.split(":")
                        start_row = int("".join(filter(str.isdigit, start_ref)))
                        end_row = int("".join(filter(str.isdigit, end_ref)))
                        return start_row, end_row
                    else:
                        # Single cell
                        row = int("".join(filter(str.isdigit, range_address)))
                        return row, row
    except Exception as e:
        logger.warning(f"Error finding range by name {name}: {str(e)}")
    return None


def get_range_by_name_in_sheet(sheet: Worksheet, name: str) -> tuple | None:
    """Find the range (start_row, end_row) of a named range in a sheet"""
    try:
        for named_range in sheet.defined_names:
            if named_range == name:
                destinations = sheet.defined_names[named_range].destinations
                for sheet_title, range_address in destinations:
                    if sheet_title == sheet.title:
                        # For named ranges that refer to entire rows or multiple rows
                        if ":" in range_address:
                            # Convert Excel range to row numbers
                            start_ref, end_ref = range_address.split(":")
                            start_row = int("".join(filter(str.isdigit, start_ref)))
                            end_row = int("".join(filter(str.isdigit, end_ref)))
                            return (start_row, end_row)
                        else:
                            # Single cell
                            row = int("".join(filter(str.isdigit, range_address)))
                            return (row, row)
    except Exception as e:
        logger.warning(f"Error finding range by name {name}: {str(e)}")
    return None


def get_cell_by_name_in_workbook(workbook: Workbook, name: str) -> Cell | None:
    """Find a cell by name within a workbook"""
    try:
        for named_range in workbook.defined_names:
            if named_range == name:
                destinations = workbook.defined_names[named_range].destinations
                for sheet_title, coordinate in destinations:
                    sheet = workbook[sheet_title]
                    if sheet_title == sheet.title:
                        return sheet[coordinate]
    except Exception as e:
        logger.warning(f"Error finding cell by name {name}: {str(e)}")
    return None


def get_cell_value_by_name(workbook: Workbook, name: str) -> str | None:
    """Get the value of a named cell"""
    try:
        cell = get_cell_by_name_in_workbook(workbook, name)
        if cell.value is not None:
            value = str(cell.value).strip()
            return value if value else None
    except Exception as e:
        logger.warning(f"Error getting cell value by name {name}: {str(e)}")
    return None


def get_cell_value_by_name_in_sheet(sheet: Worksheet, name: str) -> str | None:
    """Get the value of a named cell within a specific sheet"""
    try:
        for named_range in sheet.defined_names:
            if named_range == name:
                destinations = sheet.defined_names[named_range].destinations
                for sheet_title, coordinate in destinations:
                    if sheet_title == sheet.title:
                        cell = sheet[coordinate]
                        if cell.value is not None:
                            value = str(cell.value).strip()
                            return value if value else None
    except Exception as e:
        logger.warning(f"Error getting cell value by name {name} in sheet {sheet.title}: {str(e)}")
    return None


def get_cell_value(row, col_idx):
    """Safely get cell value from a row by column index"""
    if col_idx is None:
        return None
    try:
        cell = row[col_idx]
        if cell.value is not None:
            value = str(cell.value).strip()
            return value if value else None
        return None
    except (IndexError, AttributeError):
        return None


def get_cell_value_by_position(sheet, row_idx, col_idx):
    """Get cell value by row and column indices (0-based)"""
    try:
        cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)  # Convert to 1-based indices
        if cell.value is not None:
            value = str(cell.value).strip()
            return value if value else None
        return None
    except Exception as e:
        logger.warning(f"Error getting cell value by position ({row_idx}, {col_idx}): {str(e)}")
        return None


def import_support(workbook: Workbook) -> Optional[List[Support]]:
    """Extract support information from the Support sheet"""
    try:
        support_sheet = workbook["Support"]
        if not support_sheet:
            return None

        support_range = get_range_by_name_in_workbook(workbook, "support")
        if not support_range:
            return None

        header_row = list(support_sheet.rows)[support_range[0] - 1]
        headers = {}
        for i, cell in enumerate(header_row):
            if cell.value:
                headers[cell.value.lower()] = i

        support_channels = []
        for row_idx in range(support_range[0], support_range[1]):
            if len(list(support_sheet.rows)) < row_idx + 1:
                break
            row = list(support_sheet.rows)[row_idx]

            channel = get_cell_value(row, headers.get("channel"))
            if not channel or row_idx == support_range[0] - 1:
                continue

            support_channel = Support(
                channel=channel,
                url=get_cell_value(row, headers.get("channel url")),
                description=get_cell_value(row, headers.get("description")),
                tool=get_cell_value(row, headers.get("tool")),
                scope=get_cell_value(row, headers.get("scope")),
                invitationUrl=get_cell_value(row, headers.get("invitation url")),
            )

            support_channels.append(support_channel)
    except Exception as e:
        logger.warning(f"Error importing support: {str(e)}")
        return None

    return support_channels if support_channels else None


def import_team(workbook: Workbook) -> Optional[List[Team]]:
    """Extract team information from the Team sheet"""
    try:
        team_sheet = workbook["Team"]
        if not team_sheet:
            return None

        team_range = get_range_by_name_in_workbook(workbook, "team")
        if not team_range:
            return None

        header_row = list(team_sheet.rows)[team_range[0] - 1]
        headers = {}
        for i, cell in enumerate(header_row):
            if cell.value:
                headers[cell.value.lower()] = i

        team_members = []
        for row_idx in range(team_range[0], team_range[1]):
            if len(list(team_sheet.rows)) < row_idx + 1:
                break
            row = list(team_sheet.rows)[row_idx]

            username = get_cell_value(row, headers.get("username"))
            name = get_cell_value(row, headers.get("name"))
            role = get_cell_value(row, headers.get("role"))

            if (not (username or name or role)) or row_idx == team_range[0] - 1:
                continue

            team_member = TeamMember(
                username=username,
                name=name,
                description=get_cell_value(row, headers.get("description")),
                role=role,
                dateIn=get_cell_value(row, headers.get("date in")),
                dateOut=get_cell_value(row, headers.get("date out")),
                replacedByUsername=get_cell_value(row, headers.get("replaced by username")),
            )

            team_members.append(team_member)
    except Exception as e:
        logger.warning(f"Error importing team: {str(e)}")
        return None

    return team_members if team_members else None


def import_roles(workbook: Workbook) -> Optional[List[Role]]:
    """Extract roles information from the Roles sheet"""
    try:
        roles_sheet = workbook["Roles"]
        if not roles_sheet:
            return None

        roles_range = get_range_by_name_in_sheet(roles_sheet, "roles")
        if not roles_range:
            return None

        header_row = list(roles_sheet.rows)[roles_range[0] - 1]
        headers = {}
        for i, cell in enumerate(header_row):
            if cell.value:
                headers[cell.value.lower()] = i

        roles_list = []
        for row_idx in range(roles_range[0], roles_range[1]):
            if len(list(roles_sheet.rows)) < row_idx + 1:
                break
            row = list(roles_sheet.rows)[row_idx]

            role_name = get_cell_value(row, headers.get("role"))
            if not role_name or row_idx == roles_range[0] - 1:
                continue

            role = Role(
                role=role_name,
                description=get_cell_value(row, headers.get("description")),
                access=get_cell_value(row, headers.get("access")),
                firstLevelApprovers=get_cell_value(row, headers.get("1st level approvers")),
                secondLevelApprovers=get_cell_value(row, headers.get("2nd level approvers")),
                customProperties=None,
            )

            roles_list.append(role)
    except Exception as e:
        logger.warning(f"Error importing roles: {str(e)}")
        return None

    return roles_list if roles_list else None


def import_sla_properties(workbook: Workbook) -> Optional[List[ServiceLevelAgreementProperty]]:
    """Extract SLA properties from the SLA sheet"""
    try:
        sla_sheet = workbook["SLA"]
        if not sla_sheet:
            return None

        sla_range = get_range_by_name_in_sheet(sla_sheet, "slaProperties")
        if not sla_range:
            return None

        header_row = list(sla_sheet.rows)[sla_range[0] - 1]
        headers = {}
        for i, cell in enumerate(header_row):
            if cell.value:
                headers[cell.value.lower()] = i

        sla_properties = []
        for row_idx in range(sla_range[0], sla_range[1]):
            if len(list(sla_sheet.rows)) < row_idx + 1:
                break
            row = list(sla_sheet.rows)[row_idx]

            property_name = get_cell_value(row, headers.get("property"))
            if not property_name or row_idx == sla_range[0] - 1:
                continue

            sla_property = ServiceLevelAgreementProperty(
                property=property_name,
                value=get_cell_value(row, headers.get("value")),
                valueExt=get_cell_value(row, headers.get("extended value")),
                unit=get_cell_value(row, headers.get("unit")),
                element=get_cell_value(row, headers.get("element")),
                driver=get_cell_value(row, headers.get("driver")),
            )

            sla_properties.append(sla_property)
    except Exception as e:
        logger.warning(f"Error importing SLA properties: {str(e)}")
        return None

    return sla_properties if sla_properties else None


def import_servers(workbook) -> Optional[List[Server]]:
    """Extract server information from the Servers sheet"""
    try:
        sheet = workbook["Servers"]
        if not sheet:
            return None

        # Find the server cells
        server_cell = get_cell_by_name_in_workbook(workbook, "servers.server")
        if not server_cell:
            return None

        # Get servers (horizontally arranged in the sheet)
        servers = []
        col_idx = server_cell.column - 1  # 0-based index
        row_idx = server_cell.row - 1  # 0-based index

        index = 0
        while True:
            server_name = get_cell_value_by_position(sheet, row_idx, col_idx + index)
            if not server_name:
                break

            server = Server(
                server=server_name,
                description=get_server_cell_value(workbook, sheet, "servers.description", index),
                environment=get_server_cell_value(workbook, sheet, "servers.environment", index),
                type=get_server_cell_value(workbook, sheet, "servers.type", index),
            )

            # Get type-specific fields
            server_type = server.type
            if server_type:
                if server_type == "azure":
                    server.location = get_server_cell_value(workbook, sheet, "servers.azure.location", index)
                    server.format = get_server_cell_value(workbook, sheet, "servers.azure.format", index)
                    server.delimiter = get_server_cell_value(workbook, sheet, "servers.azure.delimiter", index)
                elif server_type == "bigquery":
                    server.project = get_server_cell_value(workbook, sheet, "servers.bigquery.project", index)
                    server.dataset = get_server_cell_value(workbook, sheet, "servers.bigquery.dataset", index)
                elif server_type == "databricks":
                    server.catalog = get_server_cell_value(workbook, sheet, "servers.databricks.catalog", index)
                    server.host = get_server_cell_value(workbook, sheet, "servers.databricks.host", index)
                    server.schema_ = get_server_cell_value(workbook, sheet, "servers.databricks.schema", index)
                elif server_type == "glue":
                    server.account = get_server_cell_value(workbook, sheet, "servers.glue.account", index)
                    server.database = get_server_cell_value(workbook, sheet, "servers.glue.database", index)
                    server.format = get_server_cell_value(workbook, sheet, "servers.glue.format", index)
                    server.location = get_server_cell_value(workbook, sheet, "servers.glue.location", index)
                elif server_type == "kafka":
                    server.format = get_server_cell_value(workbook, sheet, "servers.kafka.format", index)
                    server.host = get_server_cell_value(workbook, sheet, "servers.kafka.host", index)
                    server.topic = get_server_cell_value(workbook, sheet, "servers.kafka.topic", index)
                elif server_type == "postgres":
                    server.database = get_server_cell_value(workbook, sheet, "servers.postgres.database", index)
                    server.host = get_server_cell_value(workbook, sheet, "servers.postgres.host", index)
                    server.port = parse_integer(get_server_cell_value(workbook, sheet, "servers.postgres.port", index))
                    server.schema_ = get_server_cell_value(workbook, sheet, "servers.postgres.schema", index)
                elif server_type == "s3":
                    server.delimiter = get_server_cell_value(workbook, sheet, "servers.s3.delimiter", index)
                    server.endpointUrl = get_server_cell_value(workbook, sheet, "servers.s3.endpointUrl", index)
                    server.format = get_server_cell_value(workbook, sheet, "servers.s3.format", index)
                    server.location = get_server_cell_value(workbook, sheet, "servers.s3.location", index)
                elif server_type == "snowflake":
                    server.account = get_server_cell_value(workbook, sheet, "servers.snowflake.account", index)
                    server.database = get_server_cell_value(workbook, sheet, "servers.snowflake.database", index)
                    server.host = get_server_cell_value(workbook, sheet, "servers.snowflake.host", index)
                    server.port = parse_integer(get_server_cell_value(workbook, sheet, "servers.snowflake.port", index))
                    server.schema_ = get_server_cell_value(workbook, sheet, "servers.snowflake.schema", index)
                    server.warehouse = get_server_cell_value(workbook, sheet, "servers.snowflake.warehouse", index)
                elif server_type == "sqlserver":
                    server.database = get_server_cell_value(workbook, sheet, "servers.sqlserver.database", index)
                    server.host = get_server_cell_value(workbook, sheet, "servers.sqlserver.host", index)
                    server.port = parse_integer(get_server_cell_value(workbook, sheet, "servers.sqlserver.port", index))
                    server.schema_ = get_server_cell_value(workbook, sheet, "servers.sqlserver.schema", index)
                elif server_type == "oracle":
                    server.host = get_server_cell_value(workbook, sheet, "servers.oracle.host", index)
                    server.port = parse_integer(get_server_cell_value(workbook, sheet, "servers.oracle.port", index))
                    server.serviceName = get_server_cell_value(workbook, sheet, "servers.oracle.servicename", index)
                else:
                    # Custom server type - grab all possible fields
                    server.account = get_server_cell_value(workbook, sheet, "servers.custom.account", index)
                    server.catalog = get_server_cell_value(workbook, sheet, "servers.custom.catalog", index)
                    server.database = get_server_cell_value(workbook, sheet, "servers.custom.database", index)
                    server.dataset = get_server_cell_value(workbook, sheet, "servers.custom.dataset", index)
                    server.delimiter = get_server_cell_value(workbook, sheet, "servers.custom.delimiter", index)
                    server.endpointUrl = get_server_cell_value(workbook, sheet, "servers.custom.endpointUrl", index)
                    server.format = get_server_cell_value(workbook, sheet, "servers.custom.format", index)
                    server.host = get_server_cell_value(workbook, sheet, "servers.custom.host", index)
                    server.location = get_server_cell_value(workbook, sheet, "servers.custom.location", index)
                    server.path = get_server_cell_value(workbook, sheet, "servers.custom.path", index)
                    server.port = parse_integer(get_server_cell_value(workbook, sheet, "servers.custom.port", index))
                    server.project = get_server_cell_value(workbook, sheet, "servers.custom.project", index)
                    server.schema_ = get_server_cell_value(workbook, sheet, "servers.custom.schema", index)
                    server.stagingDir = get_server_cell_value(workbook, sheet, "servers.custom.stagingDir", index)
                    server.warehouse = get_server_cell_value(workbook, sheet, "servers.custom.warehouse", index)
                    server.region = get_server_cell_value(workbook, sheet, "servers.custom.region", index)
                    server.regionName = get_server_cell_value(workbook, sheet, "servers.custom.regionName", index)
                    server.serviceName = get_server_cell_value(workbook, sheet, "servers.custom.serviceName", index)

            servers.append(server)
            index += 1
    except Exception as e:
        logger.warning(f"Error importing servers: {str(e)}")
        return None

    return servers if servers else None


def get_server_cell_value(workbook: Workbook, sheet: Worksheet, name: str, col_offset: int):
    """Get cell value for server properties (arranged horizontally)"""
    try:
        cell = get_cell_by_name_in_workbook(workbook, name)
        if not cell:
            return None

        row = cell.row - 1  # 0-based
        col = cell.column - 1 + col_offset  # 0-based
        return get_cell_value_by_position(sheet, row, col)
    except Exception as e:
        logger.warning(f"Error getting server cell value for {name}: {str(e)}")
        return None


def import_price(workbook) -> Optional[Dict[str, Any]]:
    """Extract price information"""
    try:
        price_amount = get_cell_value_by_name(workbook, "price.priceAmount")
        price_currency = get_cell_value_by_name(workbook, "price.priceCurrency")
        price_unit = get_cell_value_by_name(workbook, "price.priceUnit")

        if not (price_amount or price_currency or price_unit):
            return None

        # Create a dictionary for price since the class doesn't seem to be directly available
        return {
            "priceAmount": price_amount,
            "priceCurrency": price_currency,
            "priceUnit": price_unit,
        }
    except Exception as e:
        logger.warning(f"Error importing price: {str(e)}")
        return None


def import_custom_properties(workbook: Workbook) -> List[CustomProperty]:
    """Extract custom properties"""
    custom_properties = []

    owner = get_cell_value_by_name(workbook, "owner")

    # Add owner as a custom property
    if owner:
        custom_properties.append(
            CustomProperty(
                property="owner",
                value=owner,
            )
        )

    try:
        # Get other custom properties
        custom_properties_sheet = workbook["Custom Properties"]
        if custom_properties_sheet:
            custom_properties_range = get_range_by_name_in_workbook(workbook, "CustomProperties")
            if custom_properties_range:
                # Skip header row
                for row_idx in range(custom_properties_range[0], custom_properties_range[1]):
                    if row_idx == custom_properties_range[0] - 1:
                        continue

                    property_name = get_cell_value_by_position(custom_properties_sheet, row_idx, 0)
                    if not property_name or property_name == "owner":
                        continue

                    property_value = get_cell_value_by_position(custom_properties_sheet, row_idx, 1)
                    parsed_value = parse_property_value(property_value)

                    custom_properties.append(
                        CustomProperty(
                            property=property_name,
                            value=parsed_value,
                        )
                    )
    except Exception as e:
        logger.warning(f"Error importing custom properties: {str(e)}")

    return custom_properties if custom_properties else None


def parse_property_value(value: str) -> Any:
    """Parse a property value into the appropriate type based on Excel values"""
    if value is None:
        return None

    # Try to convert to boolean (simple case)
    if isinstance(value, str):
        value_lower = value.lower().strip()
        if value_lower == "true":
            return True
        if value_lower == "false":
            return False

    # Try numeric conversions
    try:
        # Check if it's an integer
        if isinstance(value, str) and value.isdigit():
            return int(value)

        # Try float conversion
        float_val = float(value)
        # If it's a whole number, return as int
        if float_val.is_integer():
            return int(float_val)
        return float_val
    except (ValueError, TypeError, AttributeError):
        # If conversion fails, return original string
        return value


def import_quality(workbook: Workbook) -> Dict[str, List[DataQuality]]:
    """
    Import quality data from Quality sheet and organize by schema.property key

    Returns:
        Dictionary mapping schema.property keys to lists of DataQuality objects
    """
    try:
        quality_sheet = workbook["Quality"]
        if not quality_sheet:
            return {}
    except KeyError:
        logger.warning("Quality sheet not found")
        return {}

    try:
        quality_range = get_range_by_name_in_workbook(workbook, "quality")
        if not quality_range:
            logger.warning("Quality range not found")
            return {}

        quality_header_row_index = quality_range[0] - 1
        headers = get_headers_from_header_row(quality_sheet, quality_header_row_index)

        quality_map = {}

        for row_idx in range(quality_range[0], quality_range[1]):
            if len(list(quality_sheet.rows)) < row_idx + 1:
                break
            row = list(quality_sheet.rows)[row_idx]

            # Extract quality fields from row
            schema_name = get_cell_value(row, headers.get("schema"))
            property_name = get_cell_value(row, headers.get("property"))
            quality_type = get_cell_value(row, headers.get("quality type"))
            description = get_cell_value(row, headers.get("description"))
            rule = get_cell_value(row, headers.get("rule (library)"))
            query = get_cell_value(row, headers.get("query (sql)"))
            engine = get_cell_value(row, headers.get("quality engine (custom)"))
            implementation = get_cell_value(row, headers.get("implementation (custom)"))
            severity = get_cell_value(row, headers.get("severity"))
            scheduler = get_cell_value(row, headers.get("scheduler"))
            schedule = get_cell_value(row, headers.get("schedule"))
            threshold_operator = get_cell_value(row, headers.get("threshold operator"))
            threshold_value = get_cell_value(row, headers.get("threshold value"))

            # Skip if no schema name or insufficient quality data
            if not schema_name or (not quality_type and not description and not rule):
                continue

            # Parse threshold values based on operator
            threshold_dict = parse_threshold_values(threshold_operator, threshold_value)

            # Create DataQuality object with parsed thresholds
            quality = DataQuality(
                name=None,
                description=description,
                type=quality_type,
                rule=rule,
                unit=None,
                validValues=None,
                query=query,
                engine=engine,
                implementation=implementation,
                dimension=None,
                method=None,
                severity=severity,
                businessImpact=None,
                customProperties=None,
                authoritativeDefinitions=None,
                tags=None,
                scheduler=scheduler,
                schedule=schedule,
                **threshold_dict,  # Unpack threshold values
            )

            # Create key for mapping - use schema.property format
            key = schema_name if not property_name else f"{schema_name}.{property_name}"

            if key not in quality_map:
                quality_map[key] = []
            quality_map[key].append(quality)

    except Exception as e:
        logger.warning(f"Error importing quality: {str(e)}")
        return {}

    return quality_map


def parse_threshold_values(threshold_operator: str, threshold_value: str) -> Dict[str, Any]:
    """
    Parse threshold operator and value into DataQuality threshold fields

    Args:
        threshold_operator: The threshold operator (e.g., "mustBe", "mustBeBetween")
        threshold_value: The threshold value (string representation)

    Returns:
        Dictionary with appropriate threshold fields set
    """
    threshold_dict = {}

    if not threshold_operator or not threshold_value:
        return threshold_dict

    # Parse threshold values based on operator
    if threshold_operator in ["mustBeBetween", "mustNotBeBetween"]:
        # Parse "[value1, value2]" format
        if threshold_value.startswith("[") and threshold_value.endswith("]"):
            content = threshold_value[1:-1]  # Remove brackets
            try:
                values = [Decimal(v.strip()) for v in content.split(",") if v.strip()]
                if len(values) >= 2:
                    threshold_dict[threshold_operator] = values[:2]  # Take first two values
            except (ValueError, TypeError) as e:
                logger.warning(f"Failed to parse between values: {threshold_value}, error: {e}")
    else:
        # Single value for other operators
        try:
            # Try to parse as number
            isFraction =  "." in threshold_value
            if threshold_value.replace(".", "").replace("-", "").isdigit():
                if isFraction:
                    value = float(threshold_value)
                else:
                    value = int(threshold_value)

            threshold_dict[threshold_operator] = value

        except (ValueError, TypeError) as e:
            logger.warning(f"Failed to parse threshold value: {threshold_value}, error: {e}")

    return threshold_dict


def attach_quality_to_schemas(
    schemas: Optional[List[SchemaObject]], quality_map: Dict[str, List[DataQuality]]
) -> Optional[List[SchemaObject]]:
    """
    Attach quality attributes to schemas and their properties based on quality_map

    Args:
        schemas: List of schema objects
        quality_map: Dictionary mapping schema.property keys to quality lists

    Returns:
        List of schema objects with quality attached
    """
    if not schemas:
        return None

    updated_schemas = []

    for schema in schemas:
        schema_name = schema.name
        if not schema_name:
            updated_schemas.append(schema)
            continue

        # Get schema-level quality attributes
        schema_quality = quality_map.get(schema_name)
        if schema_quality:
            schema.quality = schema_quality

        # Attach quality to properties
        if schema.properties:
            schema.properties = attach_quality_to_properties(schema.properties, schema_name, quality_map)

        updated_schemas.append(schema)

    return updated_schemas


def attach_quality_to_properties(
    properties: List[SchemaProperty], schema_name: str, quality_map: Dict[str, List[DataQuality]], prefix: str = ""
) -> List[SchemaProperty]:
    """
    Recursively attach quality attributes to properties and nested properties

    Args:
        properties: List of property objects
        schema_name: Name of the parent schema
        quality_map: Dictionary mapping schema.property keys to quality lists
        prefix: Current property path prefix for nested properties

    Returns:
        List of property objects with quality attached
    """
    updated_properties = []

    for prop in properties:
        property_name = prop.name
        if not property_name:
            updated_properties.append(prop)
            continue

        # Build full property path
        full_property_name = f"{prefix}.{property_name}" if prefix else property_name
        quality_key = f"{schema_name}.{full_property_name}"

        # Get quality for this property
        property_quality = quality_map.get(quality_key)
        if property_quality:
            prop.quality = property_quality

        # Handle nested properties
        if prop.properties:
            prop.properties = attach_quality_to_properties(
                prop.properties, schema_name, quality_map, full_property_name
            )

        # Handle array items
        if prop.items:
            items_quality_key = f"{schema_name}.{full_property_name}.items"
            items_quality = quality_map.get(items_quality_key)
            if items_quality:
                prop.items.quality = items_quality

            # Handle nested properties in array items
            if prop.items.properties:
                prop.items.properties = attach_quality_to_properties(
                    prop.items.properties, schema_name, quality_map, f"{full_property_name}.items"
                )

        updated_properties.append(prop)

    return updated_properties


def get_headers_from_header_row(sheet: Worksheet, header_row_index: int) -> Dict[str, int]:
    """
    Get headers from the first row and map them to column indices

    Args:
        sheet: The worksheet
        header_row_index: 0-based row index of the header row

    Returns:
        Dictionary mapping header names (lowercase) to column indices
    """
    headers = {}
    try:
        header_row = list(sheet.rows)[header_row_index]
        for i, cell in enumerate(header_row):
            if cell.value:
                headers[str(cell.value).lower().strip()] = i
    except (IndexError, AttributeError) as e:
        logger.warning(f"Error getting headers from row {header_row_index}: {e}")

    return headers
