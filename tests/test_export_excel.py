import io
import os
import tempfile

import openpyxl
from open_data_contract_standard.model import OpenDataContractStandard
from typer.testing import CliRunner

from datacontract.cli import app
from datacontract.export.excel_exporter import export_to_excel_bytes
from datacontract.imports.excel_importer import import_excel_as_odcs


def test_cli_export_excel():
    """Test Excel export via CLI"""
    runner = CliRunner()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_path = tmp_file.name

    try:
        result = runner.invoke(
            app,
            [
                "export",
                "./fixtures/excel/shipments-odcs.yaml",
                "--format",
                "excel",
                "--output",
                tmp_path,
            ],
        )
        assert result.exit_code == 0
        assert os.path.exists(tmp_path)

        # Verify the file is a valid Excel file
        workbook = openpyxl.load_workbook(tmp_path)
        assert "Fundamentals" in workbook.sheetnames
        assert "Quality" in workbook.sheetnames
        assert "Schema shipments" in workbook.sheetnames
        workbook.close()

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def test_export_excel_odcs():
    """Test Excel export from ODCS object"""
    # Load the test fixture
    with open("./fixtures/excel/shipments-odcs.yaml", "r") as f:
        odcs = OpenDataContractStandard.from_string(f.read())

    # Export to Excel
    excel_bytes = export_to_excel_bytes(odcs)

    # Verify it's valid Excel data
    assert len(excel_bytes) > 0

    # Load the Excel to verify structure
    workbook = openpyxl.load_workbook(io.BytesIO(excel_bytes))

    # Check required sheets exist
    expected_sheets = [
        "Fundamentals",
        "Quality",
        "Schema shipments",
        "Support",
        "Team",
        "Roles",
        "SLA",
        "Servers",
        "Custom Properties",
    ]

    for sheet_name in expected_sheets:
        assert sheet_name in workbook.sheetnames, f"Missing sheet: {sheet_name}"

    workbook.close()


def test_excel_roundtrip():
    """Test that export then import produces equivalent data"""
    # Load original ODCS
    with open("./fixtures/excel/shipments-odcs.yaml", "r") as f:
        original_odcs = OpenDataContractStandard.from_string(f.read())

    # Export to Excel bytes
    excel_bytes = export_to_excel_bytes(original_odcs)

    # Save to temporary file for import
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file.write(excel_bytes)
        tmp_path = tmp_file.name

    try:
        # Import back from Excel
        imported_odcs = import_excel_as_odcs(tmp_path)

        assert imported_odcs.to_yaml() == original_odcs.to_yaml(), "Reimported ODCS should match original"

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
